"""
leggi_excel.py
Legge il file Excel dei risultati 7° edizione Nolimpiadi e salva tutti i dati in JSON.
Esegui: python scripts/leggi_excel.py
"""
import json, sys, os

try:
    import openpyxl
except ImportError:
    print("Installazione openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

EXCEL_PATH = r"c:\Users\info\Documents\Noli\nolimpiadi2026\Punteggi e classifiche dei gironi di qualificazione (7° edizione).xlsx"
OUT_PATH   = r"c:\Users\info\Documents\Noli\nolimpiadi2026\scripts\excel_data.json"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
result = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Salta righe completamente vuote
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            rows.append([str(c) if c is not None else "" for c in row])
    result[sheet_name] = rows
    print(f"  Foglio '{sheet_name}': {len(rows)} righe")

with open(OUT_PATH, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n✅ Dati salvati in: {OUT_PATH}")
